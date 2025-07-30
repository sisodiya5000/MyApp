# app.py
import streamlit as st
import pandas as pd
import plotly.graph_objects as go
from datetime import datetime, timedelta
import random
import requests
from io import BytesIO
from streamlit_autorefresh import st_autorefresh
from openpyxl.utils import get_column_letter

# -------------------- Page & Theme Setup --------------------
st.set_page_config(
    page_title="Price Fluctuation Dashboard",
    layout="wide",
    initial_sidebar_state="expanded",
    page_icon="📊"
)

# -------------------- Sidebar Configuration --------------------
st.sidebar.title("⚙ Configuration")
# Theme selector
theme = st.sidebar.selectbox("Theme", ["Light", "Dark"], index=0)
# Data mode and date inputs
mode = st.sidebar.selectbox("Data Mode", ["Live", "Dummy"])
# Price plot filter
price_filter = st.sidebar.selectbox("Price Plot Filter", ["Show All", "Gold Only", "Silver Only", "Gold & Silver Only"])
start_date = st.sidebar.date_input(
    "Start Date", value=datetime.today() - timedelta(days=30),
    min_value=datetime(2000, 1, 1), max_value=datetime.today()
)
end_date = st.sidebar.date_input(
    "End Date", value=datetime.today(),
    min_value=start_date, max_value=datetime.today()
)
# Auto refresh toggle
refresh = st.sidebar.checkbox("Auto-Refresh Hourly", value=True)
if refresh:
    st_autorefresh(interval=3600000, limit=None, key="auto-refresh")
st.sidebar.markdown("---")
st.sidebar.write("Built with ❤ using Streamlit")

# -------------------- CSS Styling --------------------
if theme == "Dark":
    st.markdown("""
    <style>
        .css-18e3th9 { padding-top: 2rem; }
        body, .stApp, .block-container { background-color: #1e1e1e; color: #e0e0e0; }
        .stDownloadButton>button { background-color: #444444; color: #e0e0e0; }
        .stDownloadButton>button:hover { background-color: #555555; }
        .css-1dq8tca h2 { color: #4DB6AC; }
    </style>
    """, unsafe_allow_html=True)
    plotly_template = 'plotly_dark'
else:
    st.markdown("""
    <style>
        .css-18e3th9 { padding-top: 2rem; }
        .stDownloadButton>button { background-color: #00BFFF; color: white; }
        .stDownloadButton>button:hover { background-color: #0099cc; }
        .css-1dq8tca h2 { color: #00BFFF; }
    </style>
    """, unsafe_allow_html=True)
    plotly_template = 'plotly_white'

# -------------------- Live Price Fetch --------------------
@st.cache_data(ttl=3600)
def fetch_live_prices():
    API_KEY = "992bdd9c38f2de5d1442c75e4f661b6b"
    BASE_URL = "https://api.metalpriceapi.com/v1/latest"
    params = {"api_key": API_KEY, "base": "USD", "currencies": "XAG,XAU"}
    try:
        response = requests.get(BASE_URL, params=params, timeout=10).json()
        if "error" in response:
            st.error(f"❌ API Error: {response['error']['message']}")
            return {}
        rates = response.get("rates", {})
        return {
            "Silver (USD/ton) [Live]": round(rates.get("XAG", 0) * 32015.5, 2),
            "Gold   (USD/ton) [Live]": round(rates.get("XAU", 0) * 32015.5, 2)
        }
    except Exception as e:
        st.error(f"⚠ Failed to fetch live prices: {e}")
        return {}

# -------------------- Dummy Prices --------------------
def fetch_dummy_prices():
    return {
        "Copper (USD/ton) [Dummy]": 8700,
        "Silver (USD/ton) [Dummy]": 29500,
        "Aluminum (USD/ton) [Dummy]": 2300,
        "Steel (USD/ton) [Dummy]": 600,
        "Zinc (USD/ton) [Dummy]": 2600,
        "Tin (USD/ton) [Dummy]": 27000,
        "Polysilicon (USD/kg) [Dummy]": 9.5,
        "Polymer (USD/kg) [Dummy]": 1.6,
        "Solar Glass (USD/sqm) [Dummy]": 4.2,
        "Silver Paste (USD/kg) [Dummy]": 790,
        "Solar Cell (USD/W) [Dummy]": 0.17,
        "EPE (USD/sqm) [Dummy]": 0.9,
        "Aluminum Frame (USD/ton) [Dummy]": 2500
    }

# -------------------- Fetch & Prepare Data --------------------
days = (end_date - start_date).days + 1
prices = fetch_live_prices() if mode == "Live" else fetch_dummy_prices()
# Build DataFrame
dates = [start_date + timedelta(days=i) for i in range(days)]
data = {"Date": dates}
for key, base in prices.items(): data[key] = [round(base + random.uniform(-20, 20), 2) for _ in dates]
df = pd.DataFrame(data); df['Date'] = pd.to_datetime(df['Date'])

# -------------------- Header & Metrics --------------------
st.title("📊 Price Fluctuation Dashboard")
st.markdown(f"*Last Updated:* {datetime.now():%Y-%m-%d %H:%M:%S}")
st.markdown("---")
avg = df.drop(columns='Date').mean().round(2)
cols = st.columns(len(avg))
for c, (n, v) in zip(cols, avg.items()): c.metric(label=n, value=f"{v:,.2f}")

# -------------------- Trend Chart --------------------
st.subheader("📈 Trend Over Time")
fig = go.Figure()

# Filter columns based on selection
if price_filter == "Gold Only":
    filtered_cols = [col for col in avg.index if 'Gold' in col or 'gold' in col]
elif price_filter == "Silver Only":
    filtered_cols = [col for col in avg.index if 'Silver' in col or 'silver' in col]
elif price_filter == "Gold & Silver Only":
    filtered_cols = [col for col in avg.index if 'Gold' in col or 'gold' in col or 'Silver' in col or 'silver' in col]
else:  # Show All
    filtered_cols = avg.index

for col in filtered_cols:
    fig.add_trace(go.Scatter(x=df['Date'], y=df[col], mode='lines+markers', name=col))

fig.update_layout(template=plotly_template, height=450, xaxis_title='Date', yaxis_title='Price', legend_title='Commodity')
st.plotly_chart(fig, use_container_width=True)

# -------------------- Data Table --------------------
st.subheader("📄 Data Table")
st.dataframe(df.style.format({c: "{:.2f}" for c in df.columns if c != 'Date'}))

# -------------------- Export Report (CORRECTED) --------------------
def to_excel(df, price_filter):
    # Filter DataFrame based on price filter selection
    if price_filter == "Gold Only":
        filtered_cols = ['Date'] + [col for col in df.columns if col != 'Date' and ('Gold' in col or 'gold' in col)]
    elif price_filter == "Silver Only":
        filtered_cols = ['Date'] + [col for col in df.columns if col != 'Date' and ('Silver' in col or 'silver' in col)]
    elif price_filter == "Gold & Silver Only":
        filtered_cols = ['Date'] + [col for col in df.columns if col != 'Date' and ('Gold' in col or 'gold' in col or 'Silver' in col or 'silver' in col)]
    else:  # Show All
        filtered_cols = df.columns.tolist()
    
    # Create filtered dataframe
    df_filtered = df[filtered_cols]
    tr = df_filtered.copy()
    
    if days <= 7:
        sheet = 'Daily'
    elif days <= 90:
        tr['Week'] = tr['Date'].dt.to_period('W').apply(lambda r: r.start_time)
        # Group by week and calculate means for price columns only (exclude Date column)
        price_columns = [col for col in tr.columns if col not in ['Date', 'Week']]
        tr_grouped = tr.groupby('Week')[price_columns].mean().reset_index()
        tr_grouped = tr_grouped.rename(columns={'Week':'Date'})
        tr = tr_grouped
        sheet = 'Weekly'
    else:
        # Fixed monthly date range - now shows start and end dates of each month
        tr['Month'] = tr['Date'].dt.to_period('M')
        # Group by month and calculate means for price columns only (exclude Date column)
        price_columns = [col for col in tr.columns if col not in ['Date', 'Month']]
        tr_grouped = tr.groupby('Month')[price_columns].mean().reset_index()
        # Add start and end date columns for each month
        tr_grouped['Start_Date'] = tr_grouped['Month'].apply(lambda x: x.start_time)
        tr_grouped['End_Date'] = tr_grouped['Month'].apply(lambda x: x.end_time)
        # Keep only Start_Date, End_Date, and price columns
        tr = tr_grouped[['Start_Date', 'End_Date'] + price_columns]
        sheet = 'Monthly'
    
    # Calculate averages for filtered data only
    avg_filtered = df_filtered.drop(columns='Date').mean().round(2)
    
    buf = BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as w:
        df_filtered.to_excel(w, sheet_name='Raw_Data', index=False)
        avg_filtered.to_frame('Average').to_excel(w, sheet_name='Averages')
        tr.to_excel(w, sheet_name=sheet, index=False)
        for s in w.book.sheetnames:
            ws = w.book[s]
            for i, col in enumerate(ws.columns, 1): 
                ws.column_dimensions[get_column_letter(i)].width = max(len(str(cell.value)) for cell in col)+2
    buf.seek(0)
    return buf

excel_buf = to_excel(df, price_filter)
st.download_button(label="⬇ Download Report", data=excel_buf, file_name=f"report_{start_date:%Y%m%d}_{end_date:%Y%m%d}.xlsx", mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')